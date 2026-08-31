"""
从教务处导出的两份 Excel（全校课程清单/全校总课表）里只提取"专业选修"类别的行，
存成干净的 JSON，供后续按课程名匹配到培养方案数据、推断开课学期用。

"全校课程清单（仅供选课用）.xlsx"（sheet1）有明确的"课程分类"列，可以直接筛
"专业选修"。
"全校总课表.xlsx"（Sheet1）没有可靠的类别列，用前者筛出的课程编码集合去匹配
后者的"课程编号"列（两份文件用的是同一套教务系统编码，不是培养方案 JSON 里
DZ.../A20...M 那套编码，需要后续按课程名对应）。

用法：
    python3 scripts/extract_elective_schedule.py <课程清单.xlsx> <总课表.xlsx>
"""
import argparse
import json
import os
import sys

import openpyxl

try:
    import xlrd
except ImportError:  # 只有处理旧版 .xls 时才需要
    xlrd = None

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "schedule_raw")


def extract_course_list(path):
    """全校课程清单（仅供选课用）.xlsx -> 专业选修行"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["sheet1"]
    title = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[4] != "专业选修":
            continue
        rows.append(
            {
                "code": row[2],
                "name": row[3],
                "category": row[4],
                "teacher": row[5],
                "class_name": row[8],
                "college": row[10],
                "credits": row[11],
                "total_hours": row[12],
                "exam_type": row[17],
                "class_size": row[18],
                "schedule_text": row[20],
            }
        )
    return title, rows


def extract_semester_course_list(path):
    """
    形如"24春全校课表21日查看.xlsx"的单学期课表 -> 专业选修行。
    列比"课程清单（仅供选课用）"少（没有考核方式/校区），按实际表头取。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["sheet1"]
    title = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[3] != "专业选修":
            continue
        rows.append(
            {
                "code": row[1],
                "name": row[2],
                "category": row[3],
                "teacher": row[4],
                "class_name": row[5],
                "college": row[6],
                "credits": row[7],
                "total_hours": row[8],
                "class_size": row[9],
                "schedule_text": row[12],
            }
        )
    return title, rows


def extract_course_list_xls(path):
    """
    形如"2026-2027-1课程清单 (仅供选课参考).xls"的旧版 Excel 课程清单 -> 专业选修行。
    跟"全校课程清单（仅供选课用）.xlsx"结构基本一致（有独立的"课程分类"列），只是
    中间多插了一列"学生院系"，列序整体往后挪了一位，且是老版 .xls 格式要用 xlrd 读。
    """
    if xlrd is None:
        raise RuntimeError("读取 .xls 需要 xlrd，先 pip install xlrd")
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    title = sh.row_values(0)[0]

    rows = []
    for r in range(2, sh.nrows):
        row = sh.row_values(r)
        if row[5] != "专业选修":
            continue
        rows.append(
            {
                "code": row[3],
                "name": row[4],
                "category": row[5],
                "teacher": row[6],
                "class_name": row[7],
                "college": row[9],
                "credits": row[10],
                "total_hours": row[11],
                "exam_type": row[16],
                "class_size": row[17],
                "schedule_text": row[18],
            }
        )
    return title, rows


def extract_master_schedule(path, known_codes):
    """全校总课表.xlsx -> 课程编号命中 known_codes 的行（没有独立类别列，靠编码匹配）"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    title = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[4]
        if code not in known_codes:
            continue
        rows.append(
            {
                "code": code,
                "name": row[5],
                "campus": row[1],
                "college": row[2],
                "teacher": row[3],
                "class_name": row[7],
                "credits": row[8],
                "total_hours": row[9],
                "location": row[10],
                "open_type": row[18],
            }
        )
    return title, rows


def main():
    parser = argparse.ArgumentParser(description="提取全校课表里的专业选修行")
    parser.add_argument("--single", metavar="XLSX", help="单学期课表（如'24春全校课表'），有独立课程分类列")
    parser.add_argument(
        "--single-xls",
        metavar="XLS",
        help="旧版 .xls 格式的单学期课程清单（如'2026-2027-1课程清单'），有独立课程分类列",
    )
    parser.add_argument("--out-name", metavar="NAME", help="--single/--single-xls 模式下的输出文件名（不含.json）")
    parser.add_argument("course_list_xlsx", nargs="?", help="全校课程清单（仅供选课用）.xlsx")
    parser.add_argument("master_schedule_xlsx", nargs="?", help="全校总课表.xlsx")
    args = parser.parse_args()

    os.makedirs(OUT_DIR, exist_ok=True)

    if args.single or args.single_xls:
        if args.single_xls:
            title, rows = extract_course_list_xls(args.single_xls)
        else:
            title, rows = extract_semester_course_list(args.single)
        out_name = args.out_name or "single_semester_elective"
        out_path = os.path.join(OUT_DIR, f"{out_name}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump({"source_title": title, "rows": rows}, f, ensure_ascii=False, indent=2)
        codes = {r["code"] for r in rows}
        print(f"[{title}] 专业选修唯一课程编码：{len(codes)} 个，共 {len(rows)} 行 -> {out_path}")
        return

    if not args.course_list_xlsx or not args.master_schedule_xlsx:
        parser.error("需要提供 course_list_xlsx 和 master_schedule_xlsx，或使用 --single")

    title1, list_rows = extract_course_list(args.course_list_xlsx)
    codes = {r["code"] for r in list_rows}
    title2, schedule_rows = extract_master_schedule(args.master_schedule_xlsx, codes)

    out1 = os.path.join(OUT_DIR, "course_list_elective.json")
    with open(out1, "w", encoding="utf-8") as f:
        json.dump({"source_title": title1, "rows": list_rows}, f, ensure_ascii=False, indent=2)

    out2 = os.path.join(OUT_DIR, "master_schedule_elective.json")
    with open(out2, "w", encoding="utf-8") as f:
        json.dump({"source_title": title2, "rows": schedule_rows}, f, ensure_ascii=False, indent=2)

    print(f"[{title1}] 专业选修唯一课程编码：{len(codes)} 个，共 {len(list_rows)} 行 -> {out1}")
    print(f"[{title2}] 命中编码的开课行：{len(schedule_rows)} 行 -> {out2}")

    unmatched = len({r['code'] for r in list_rows}) - len({r['code'] for r in schedule_rows})
    print(f"注意：课程清单里有编码在总课表里没找到对应开课行的，约 {unmatched} 个"
          "（可能是该学期停开，或该课程在总课表所属学期不是这两份文件对应的学期）。")


if __name__ == "__main__":
    sys.exit(main())
